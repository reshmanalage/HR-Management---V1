"""
Bulk employee import via Excel (.xlsx).

Speed strategy:
  - openpyxl read_only=True  → skips cell-style tree, ~3x faster parse
  - All lookups fetched as raw tuples (no ORM object hydration, no lazy loads)
  - Bulk-create new depts/desigs in one INSERT each
  - Single multi-row INSERT for all valid employees
  - Related records (address, bank, statutory) bulk-inserted after employee IDs are fetched
  - Row-by-row fallback only if bulk INSERT fails, with per-row error capture
  - Failed rows packaged as base64 xlsx for re-upload
"""
from __future__ import annotations

import base64
import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import select, insert as sa_insert, func
from sqlalchemy.orm import Session

from app.models.department import Department
from app.models.designation import Designation
from app.models.employee import Employee, EmployeeStatus, EmploymentType, BloodGroup, MaritalStatus
from app.models.employee_address import EmployeeAddress
from app.models.employee_bank_account import EmployeeBankAccount
from app.models.employee_statutory import EmployeeStatutory

log = logging.getLogger(__name__)

# ── Column definitions ─────────────────────────────────────────────────────────
# (field_key, header_label, sample_value)

COLUMNS = [
    # ── Core (keep positions 1-17 identical to previous version) ──────────────
    ("employee_code",         "Employee Code",          "EMP0001 (auto if blank)"),
    ("first_name",            "First Name *",           "John"),
    ("middle_name",           "Middle Name",            ""),
    ("last_name",             "Last Name",              "Doe"),
    ("gender",                "Gender",                 "male / female / other"),
    ("date_of_birth",         "Date of Birth",          "1990-01-25"),
    ("personal_email",        "Personal Email",         "john@gmail.com"),
    ("company_email",         "Company Email",          "john@company.com"),
    ("mobile_number",         "Mobile Number",          "9876543210"),
    ("department",            "Department",             "Engineering"),
    ("designation",           "Designation",            "Software Engineer"),
    ("employment_type",       "Employment Type",        "permanent / probation / contract / intern / part_time / consultant"),
    ("employee_status",       "Employee Status",        "active / probation / notice_period / inactive / terminated"),
    ("date_of_joining",       "Date of Joining *",      "2024-01-01"),
    ("branch",                "Branch",                 "Mumbai"),
    ("location",              "Location",               "WFH / Office"),
    ("grade",                 "Grade",                  "L2"),
    # ── Personal (new) ─────────────────────────────────────────────────────────
    ("biometric_code",        "Biometric Code",         "73"),
    ("blood_group",           "Blood Group",            "A+ / B+ / O+ / AB+ / A- / B- / O- / AB-"),
    ("marital_status",        "Marital Status",         "single / married / divorced / widowed"),
    ("nationality",           "Nationality",            "Indian"),
    ("alternate_mobile",      "Alternate Mobile",       "9876543211"),
    # ── Employment (new) ───────────────────────────────────────────────────────
    ("confirmation_date",     "Confirmation Date",      "2024-07-01"),
    ("cost_center",           "Cost Center",            "CC-001"),
    ("reporting_manager_code","Reporting Manager Code", "EMP0010"),
    # ── Address ────────────────────────────────────────────────────────────────
    ("address_line_1",        "Address Line 1",         "123 Main Street"),
    ("city",                  "City",                   "Mumbai"),
    ("state",                 "State",                  "Maharashtra"),
    ("pincode",               "Pincode",                "400001"),
    # ── Bank ───────────────────────────────────────────────────────────────────
    ("bank_name",             "Bank Name",              "HDFC Bank"),
    ("account_number",        "Account Number",         "12345678901234"),
    ("ifsc_code",             "IFSC Code",              "HDFC0001234"),
    ("account_holder_name",   "Account Holder Name",    "John Doe"),
    ("account_type",          "Account Type",           "savings / current / salary"),
    # ── Statutory ──────────────────────────────────────────────────────────────
    ("pan_number",            "PAN Number",             "ABCDE1234F"),
    ("aadhaar_number",        "Aadhaar Number",         "123456789012"),
    ("uan_number",            "UAN Number",             "100123456789"),
    ("esic_ip_number",        "ESIC IP Number",         "1234567890"),
]

# Column index constants (1-based, matching COLUMNS list order)
C_EMP_CODE          = 1
C_FIRST_NAME        = 2
C_MIDDLE_NAME       = 3
C_LAST_NAME         = 4
C_GENDER            = 5
C_DOB               = 6
C_PERSONAL_EMAIL    = 7
C_COMPANY_EMAIL     = 8
C_MOBILE            = 9
C_DEPARTMENT        = 10
C_DESIGNATION       = 11
C_EMP_TYPE          = 12
C_EMP_STATUS        = 13
C_DOJ               = 14
C_BRANCH            = 15
C_LOCATION          = 16
C_GRADE             = 17
C_BIOMETRIC         = 18
C_BLOOD_GROUP       = 19
C_MARITAL_STATUS    = 20
C_NATIONALITY       = 21
C_ALT_MOBILE        = 22
C_CONFIRM_DATE      = 23
C_COST_CENTER       = 24
C_MANAGER_CODE      = 25
C_ADDRESS_LINE1     = 26
C_CITY              = 27
C_STATE             = 28
C_PINCODE           = 29
C_BANK_NAME         = 30
C_ACCOUNT_NUMBER    = 31
C_IFSC              = 32
C_ACCOUNT_HOLDER    = 33
C_ACCOUNT_TYPE      = 34
C_PAN               = 35
C_AADHAAR           = 36
C_UAN               = 37
C_ESIC              = 38

NCOLS = len(COLUMNS)

# Section header rows in the template for readability
SECTION_ROWS = {
    1:  ("Core Details",       "4F46E5"),  # indigo
    18: ("Personal Details",   "0891B2"),  # cyan
    23: ("Employment Details", "059669"),  # emerald
    26: ("Address",            "D97706"),  # amber
    30: ("Bank Details",       "7C3AED"),  # violet
    35: ("Statutory Details",  "DC2626"),  # red
}

HEADER_FILL    = PatternFill("solid", fgColor="4F46E5")
HEADER_FONT    = Font(color="FFFFFF", bold=True)
SAMPLE_FILL    = PatternFill("solid", fgColor="EEF2FF")
ERROR_FILL     = PatternFill("solid", fgColor="FEE2E2")
ERROR_HDR_FILL = PatternFill("solid", fgColor="DC2626")

SECTION_COL_FILLS = {
    range(1, 18):  PatternFill("solid", fgColor="EEF2FF"),  # indigo tint
    range(18, 23): PatternFill("solid", fgColor="E0F9FF"),  # cyan tint
    range(23, 26): PatternFill("solid", fgColor="ECFDF5"),  # emerald tint
    range(26, 30): PatternFill("solid", fgColor="FFFBEB"),  # amber tint
    range(30, 35): PatternFill("solid", fgColor="F5F3FF"),  # violet tint
    range(35, 39): PatternFill("solid", fgColor="FFF1F2"),  # red tint
}

COL_WIDTHS = [
    18, 15, 15, 15, 10, 14, 25, 25, 14, 20, 22, 20, 16, 14, 14, 14, 10,  # 1-17
    14, 12, 14, 14, 16,                                                    # 18-22
    14, 14, 22,                                                            # 23-25
    30, 16, 16, 10,                                                        # 26-29
    18, 18, 14, 22, 12,                                                    # 30-34
    14, 14, 14, 14,                                                        # 35-38
]

_GENDER_MAP = {"male": "male", "m": "male", "female": "female", "f": "female", "other": "other", "o": "other"}

_BLOOD_GROUP_MAP = {
    "a+": "A+", "a-": "A-", "b+": "B+", "b-": "B-",
    "ab+": "AB+", "ab-": "AB-", "o+": "O+", "o-": "O-",
}

_MARITAL_MAP = {
    "single": "single", "married": "married",
    "divorced": "divorced", "widowed": "widowed", "other": "other",
}

_ACCOUNT_TYPE_MAP = {"savings": "savings", "current": "current", "salary": "salary"}


# ── template ──────────────────────────────────────────────────────────────────

def generate_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Section header row (row 1) — spans columns per section
    section_starts = sorted(SECTION_ROWS.keys())
    for i, start_col in enumerate(section_starts):
        label, color = SECTION_ROWS[start_col]
        end_col = section_starts[i + 1] - 1 if i + 1 < len(section_starts) else NCOLS
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col, value=label)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Field header row (row 2)
    for col_idx, (_, header, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        fill = next((f for r, f in SECTION_COL_FILLS.items() if col_idx in r), SAMPLE_FILL)
        cell.fill = PatternFill("solid", fgColor=fill.fgColor)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Sample row (row 3)
    for col_idx, (_, _, sample) in enumerate(COLUMNS, start=1):
        fill = next((f for r, f in SECTION_COL_FILLS.items() if col_idx in r), SAMPLE_FILL)
        cell = ws.cell(row=3, column=col_idx, value=sample)
        cell.fill = PatternFill("solid", fgColor=fill.fgColor)

    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── result types ──────────────────────────────────────────────────────────────

@dataclass
class RowResult:
    row: int
    status: str
    employee_code: str = ""
    name: str = ""
    error: str = ""


@dataclass
class BulkImportResult:
    total: int = 0
    success: int = 0
    failed: int = 0
    rows: list[RowResult] = field(default_factory=list)
    failed_rows_xlsx_b64: str = ""


# ── helpers ───────────────────────────────────────────────────────────────────

def _s(raw: tuple, col: int) -> str:
    """Return cell value at 1-based column index as stripped string."""
    v = raw[col - 1] if col - 1 < len(raw) else None
    return str(v).strip() if v is not None else ""


def _parse_date(raw_val) -> date | None:
    if raw_val is None:
        return None
    if isinstance(raw_val, datetime):
        return raw_val.date()
    if isinstance(raw_val, date):
        return raw_val
    s = str(raw_val).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _is_empty_row(raw: tuple) -> bool:
    return all(v is None or str(v).strip() == "" for v in raw)


def _mobile(raw_val: str) -> str | None:
    if not raw_val:
        return None
    clean = re.split(r"[/|,\\]", raw_val)[0].strip()
    return clean[:20] or None


# ── failed-rows Excel ─────────────────────────────────────────────────────────

def _build_failed_xlsx(failed: list[tuple[int, tuple, str]]) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Failed Rows"

    for col_idx, (_, header, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    err_col = NCOLS + 1
    err_hdr = ws.cell(row=1, column=err_col, value="Error")
    err_hdr.fill = ERROR_HDR_FILL
    err_hdr.font = Font(color="FFFFFF", bold=True)

    for out_row, (_, values, error) in enumerate(failed, start=2):
        for col_idx in range(1, NCOLS + 1):
            v = values[col_idx - 1] if col_idx - 1 < len(values) else None
            ws.cell(row=out_row, column=col_idx, value=v).fill = ERROR_FILL
        ws.cell(row=out_row, column=err_col, value=error).fill = ERROR_FILL

    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions[get_column_letter(err_col)].width = 45
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()


# ── main processor ────────────────────────────────────────────────────────────

def process_upload(db: Session, file_bytes: bytes, created_by: int) -> BulkImportResult:
    result = BulkImportResult()

    # ── Step 1: pre-load lookups ──────────────────────────────────────────────
    log.info("bulk_import: loading lookups")

    existing_codes: set[str] = set(
        row[0] for row in db.execute(
            select(Employee.employee_code).where(Employee.employee_code.isnot(None))
        )
    )
    existing_emails: set[str] = set(
        row[0] for row in db.execute(
            select(Employee.company_email).where(Employee.company_email.isnot(None))
        )
    )
    existing_biometrics: set[str] = set(
        row[0] for row in db.execute(
            select(Employee.biometric_code).where(Employee.biometric_code.isnot(None))
        )
    )
    dept_map: dict[str, int] = {
        row[1].lower(): row[0]
        for row in db.execute(select(Department.id, Department.name))
    }
    desig_map: dict[str, int] = {
        row[1].lower(): row[0]
        for row in db.execute(select(Designation.id, Designation.title))
    }
    # Existing employee code → id (for reporting manager lookup)
    emp_code_to_id: dict[str, int] = {
        row[0]: row[1]
        for row in db.execute(select(Employee.employee_code, Employee.id))
        if row[0]
    }
    max_emp_id: int = db.scalar(select(func.max(Employee.id))) or 0

    log.info(
        "bulk_import: lookups ready — codes=%d emails=%d depts=%d desigs=%d",
        len(existing_codes), len(existing_emails), len(dept_map), len(desig_map),
    )

    # ── Step 2: parse Excel ───────────────────────────────────────────────────
    log.info("bulk_import: parsing Excel (%d bytes)", len(file_bytes))
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active

    all_rows: list[tuple[int, tuple]] = []
    consecutive_empty = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = tuple(row[:NCOLS])
        if _is_empty_row(raw):
            consecutive_empty += 1
            if consecutive_empty >= 10:
                break
        else:
            consecutive_empty = 0
            # Skip section/header rows: if column 2 (first_name) equals the label text, skip
            if _s(raw, C_FIRST_NAME).lower() in ("first name *", "first name", "sample"):
                continue
            all_rows.append((row_idx, raw))

    wb.close()
    result.total = len(all_rows)
    log.info("bulk_import: %d data rows found", result.total)

    if not all_rows:
        return result

    # ── Step 3: bulk-create new depts / desigs ────────────────────────────────
    needed_depts = {
        _s(raw, C_DEPARTMENT).lower() for _, raw in all_rows
        if _s(raw, C_DEPARTMENT) and _s(raw, C_DEPARTMENT).lower() not in dept_map
    }
    needed_desigs = {
        _s(raw, C_DESIGNATION).lower() for _, raw in all_rows
        if _s(raw, C_DESIGNATION) and _s(raw, C_DESIGNATION).lower() not in desig_map
    }

    if needed_depts:
        db.execute(sa_insert(Department), [{"name": n.title()} for n in needed_depts])
        db.flush()
        for row in db.execute(
            select(Department.id, Department.name).where(
                Department.name.in_([n.title() for n in needed_depts])
            )
        ):
            dept_map[row[1].lower()] = row[0]

    if needed_desigs:
        db.execute(sa_insert(Designation), [{"title": t.title()} for t in needed_desigs])
        db.flush()
        for row in db.execute(
            select(Designation.id, Designation.title).where(
                Designation.title.in_([t.title() for t in needed_desigs])
            )
        ):
            desig_map[row[1].lower()] = row[0]

    # ── Step 4: validate every row in memory ─────────────────────────────────
    log.info("bulk_import: validating %d rows", len(all_rows))
    valid_payloads: list[tuple[int, dict, str, str, tuple]] = []  # (row_idx, emp_dict, code, name, raw)
    failed_raw:     list[tuple[int, tuple, str]] = []
    used_codes:     set[str] = set()
    used_emails:    set[str] = set()
    used_biometrics: set[str] = set()
    auto_id = max_emp_id

    for row_idx, raw in all_rows:
        first_name = _s(raw, C_FIRST_NAME)
        last_name  = _s(raw, C_LAST_NAME)
        full_name  = f"{first_name} {last_name}".strip()

        try:
            if not first_name:
                raise ValueError("First name is required")

            doj = _parse_date(raw[C_DOJ - 1] if len(raw) >= C_DOJ else None)
            if not doj:
                raise ValueError("Date of joining is required (format: YYYY-MM-DD)")

            # Employee code
            emp_code_raw = _s(raw, C_EMP_CODE)
            if emp_code_raw:
                if emp_code_raw in existing_codes or emp_code_raw in used_codes:
                    raise ValueError(f"Employee code '{emp_code_raw}' already exists")
                emp_code = emp_code_raw
            else:
                auto_id += 1
                emp_code = f"EMP{auto_id:04d}"
                while emp_code in existing_codes or emp_code in used_codes:
                    auto_id += 1
                    emp_code = f"EMP{auto_id:04d}"
            used_codes.add(emp_code)

            # Company email uniqueness
            company_email: str | None = _s(raw, C_COMPANY_EMAIL) or None
            if company_email:
                if company_email in existing_emails or company_email in used_emails:
                    raise ValueError(f"Company email '{company_email}' already in use")
                used_emails.add(company_email)

            # Biometric code uniqueness
            biometric_code: str | None = _s(raw, C_BIOMETRIC) or None
            if biometric_code:
                if biometric_code in existing_biometrics or biometric_code in used_biometrics:
                    raise ValueError(f"Biometric code '{biometric_code}' already in use")
                used_biometrics.add(biometric_code)

            # Gender
            gender = _GENDER_MAP.get(_s(raw, C_GENDER).lower())

            # Blood group
            bg_raw = _s(raw, C_BLOOD_GROUP).strip()
            blood_group = _BLOOD_GROUP_MAP.get(bg_raw.lower())

            # Marital status
            marital_status = _MARITAL_MAP.get(_s(raw, C_MARITAL_STATUS).strip().lower())

            # Dates
            dob            = _parse_date(raw[C_DOB - 1]          if len(raw) >= C_DOB          else None)
            confirm_date   = _parse_date(raw[C_CONFIRM_DATE - 1] if len(raw) >= C_CONFIRM_DATE else None)

            # Employment type
            et_raw = _s(raw, C_EMP_TYPE).strip().lower().replace(" ", "_")
            try:
                employment_type: str | None = EmploymentType(et_raw).value if et_raw else None
            except ValueError:
                employment_type = None

            # Employee status
            es_raw = _s(raw, C_EMP_STATUS).strip().lower().replace(" ", "_")
            try:
                employee_status = EmployeeStatus(es_raw).value if es_raw else EmployeeStatus.ACTIVE.value
            except ValueError:
                employee_status = EmployeeStatus.ACTIVE.value

            # Reporting manager (lookup existing employees; new batch mates resolved after INSERT)
            manager_code = _s(raw, C_MANAGER_CODE) or None
            reporting_manager_id = emp_code_to_id.get(manager_code) if manager_code else None

            valid_payloads.append((row_idx, dict(
                employee_code=emp_code,
                biometric_code=biometric_code,
                first_name=first_name,
                middle_name=_s(raw, C_MIDDLE_NAME) or None,
                last_name=last_name or "",
                gender=gender,
                date_of_birth=dob,
                blood_group=blood_group,
                marital_status=marital_status,
                nationality=_s(raw, C_NATIONALITY) or None,
                personal_email=_s(raw, C_PERSONAL_EMAIL) or None,
                company_email=company_email,
                mobile_number=_mobile(_s(raw, C_MOBILE)),
                alternate_mobile=_mobile(_s(raw, C_ALT_MOBILE)),
                department_id=dept_map.get(_s(raw, C_DEPARTMENT).lower()) if _s(raw, C_DEPARTMENT) else None,
                designation_id=desig_map.get(_s(raw, C_DESIGNATION).lower()) if _s(raw, C_DESIGNATION) else None,
                employment_type=employment_type,
                employee_status=employee_status,
                date_of_joining=doj,
                confirmation_date=confirm_date,
                branch=_s(raw, C_BRANCH) or None,
                location=_s(raw, C_LOCATION) or None,
                grade=_s(raw, C_GRADE) or None,
                cost_center=_s(raw, C_COST_CENTER) or None,
                reporting_manager_id=reporting_manager_id,
                created_by=created_by,
                is_active=True,
            ), emp_code, full_name, raw))

        except Exception as exc:
            result.failed += 1
            result.rows.append(RowResult(row=row_idx, status="error", name=full_name, error=str(exc)))
            failed_raw.append((row_idx, raw, str(exc)))

    log.info("bulk_import: %d valid, %d invalid", len(valid_payloads), len(failed_raw))

    if not valid_payloads:
        if failed_raw:
            result.failed_rows_xlsx_b64 = _build_failed_xlsx(failed_raw)
        result.rows.sort(key=lambda r: r.row)
        return result

    # ── Step 5: bulk INSERT employees ─────────────────────────────────────────
    try:
        log.info("bulk_import: executing bulk INSERT for %d employees", len(valid_payloads))
        db.execute(sa_insert(Employee), [p for _, p, _, _, _ in valid_payloads])
        db.flush()  # flush to get IDs without committing yet
        log.info("bulk_import: bulk INSERT flushed")

        # Fetch newly created IDs by employee code
        new_codes = [code for _, _, code, _, _ in valid_payloads]
        fresh_map: dict[str, int] = {
            row[0]: row[1]
            for row in db.execute(
                select(Employee.employee_code, Employee.id).where(
                    Employee.employee_code.in_(new_codes)
                )
            )
        }

        # Merge into running emp_code_to_id (for manager resolution within same batch)
        emp_code_to_id.update(fresh_map)

        # Update reporting_manager_id for intra-batch references
        for _, payload, _, _, raw in valid_payloads:
            manager_code = _s(raw, C_MANAGER_CODE) or None
            if manager_code and payload["reporting_manager_id"] is None:
                mid = emp_code_to_id.get(manager_code)
                if mid:
                    emp_id = fresh_map.get(payload["employee_code"])
                    if emp_id:
                        db.execute(
                            Employee.__table__.update()
                            .where(Employee.id == emp_id)
                            .values(reporting_manager_id=mid)
                        )

        # ── Build related records ──────────────────────────────────────────────
        addr_payloads  = []
        bank_payloads  = []
        stat_payloads  = []

        for _, payload, emp_code, _, raw in valid_payloads:
            emp_id = fresh_map.get(emp_code)
            if not emp_id:
                continue

            # Address — insert if any field present
            addr_line1 = _s(raw, C_ADDRESS_LINE1)
            city       = _s(raw, C_CITY)
            state      = _s(raw, C_STATE)
            pincode    = _s(raw, C_PINCODE)
            if any([addr_line1, city, state, pincode]):
                addr_payloads.append(dict(
                    employee_id=emp_id,
                    address_type="current",
                    address_line_1=addr_line1 or None,
                    city=city or None,
                    state=state or None,
                    postal_code=pincode or None,
                    country="India",
                ))

            # Bank — insert if account number + bank name present
            bank_name       = _s(raw, C_BANK_NAME)
            account_number  = _s(raw, C_ACCOUNT_NUMBER)
            ifsc            = _s(raw, C_IFSC)
            acct_holder     = _s(raw, C_ACCOUNT_HOLDER) or f"{payload['first_name']} {payload['last_name']}".strip()
            acct_type_raw   = _s(raw, C_ACCOUNT_TYPE).strip().lower()
            acct_type       = _ACCOUNT_TYPE_MAP.get(acct_type_raw, "savings")
            if bank_name and account_number:
                bank_payloads.append(dict(
                    employee_id=emp_id,
                    bank_name=bank_name,
                    account_number=account_number,
                    ifsc_code=ifsc or "UNKNOWN",
                    branch_name=None,
                    account_holder_name=acct_holder,
                    account_type=acct_type,
                    is_primary=True,
                    is_verified=False,
                ))

            # Statutory — insert if any statutory field present
            pan      = _s(raw, C_PAN)
            aadhaar  = _s(raw, C_AADHAAR)
            uan      = _s(raw, C_UAN)
            esic     = _s(raw, C_ESIC)
            if any([pan, aadhaar, uan, esic]):
                stat_payloads.append(dict(
                    employee_id=emp_id,
                    pan_number=pan or None,
                    aadhaar_number=aadhaar or None,
                    uan_number=uan or None,
                    esic_ip_number=esic or None,
                    pf_eligible=True,
                    esic_eligible=True,
                ))

        if addr_payloads:
            db.execute(sa_insert(EmployeeAddress), addr_payloads)
        if bank_payloads:
            db.execute(sa_insert(EmployeeBankAccount), bank_payloads)
        if stat_payloads:
            db.execute(sa_insert(EmployeeStatutory), stat_payloads)

        db.commit()
        log.info(
            "bulk_import: committed — employees=%d addresses=%d banks=%d statutory=%d",
            len(valid_payloads), len(addr_payloads), len(bank_payloads), len(stat_payloads),
        )

        for row_idx, _, emp_code, name, _ in valid_payloads:
            result.success += 1
            result.rows.append(RowResult(row=row_idx, status="success", employee_code=emp_code, name=name))

    except Exception as bulk_exc:
        log.error("bulk_import: bulk INSERT failed (%s), falling back to row-by-row", bulk_exc)
        db.rollback()

        for row_idx, payload, emp_code, name, raw in valid_payloads:
            try:
                db.execute(sa_insert(Employee), [payload])
                db.flush()
                emp_id_row = db.execute(
                    select(Employee.id).where(Employee.employee_code == emp_code)
                ).scalar()

                if emp_id_row:
                    addr_line1 = _s(raw, C_ADDRESS_LINE1)
                    city = _s(raw, C_CITY)
                    state = _s(raw, C_STATE)
                    pincode = _s(raw, C_PINCODE)
                    if any([addr_line1, city, state, pincode]):
                        db.execute(sa_insert(EmployeeAddress), [dict(
                            employee_id=emp_id_row, address_type="current",
                            address_line_1=addr_line1 or None, city=city or None,
                            state=state or None, postal_code=pincode or None, country="India",
                        )])

                    bank_name = _s(raw, C_BANK_NAME)
                    account_number = _s(raw, C_ACCOUNT_NUMBER)
                    if bank_name and account_number:
                        acct_type = _ACCOUNT_TYPE_MAP.get(_s(raw, C_ACCOUNT_TYPE).lower(), "savings")
                        db.execute(sa_insert(EmployeeBankAccount), [dict(
                            employee_id=emp_id_row,
                            bank_name=bank_name,
                            account_number=account_number,
                            ifsc_code=_s(raw, C_IFSC) or "UNKNOWN",
                            account_holder_name=_s(raw, C_ACCOUNT_HOLDER) or name,
                            account_type=acct_type,
                            is_primary=True, is_verified=False,
                        )])

                    pan = _s(raw, C_PAN)
                    aadhaar = _s(raw, C_AADHAAR)
                    uan = _s(raw, C_UAN)
                    esic = _s(raw, C_ESIC)
                    if any([pan, aadhaar, uan, esic]):
                        db.execute(sa_insert(EmployeeStatutory), [dict(
                            employee_id=emp_id_row,
                            pan_number=pan or None, aadhaar_number=aadhaar or None,
                            uan_number=uan or None, esic_ip_number=esic or None,
                            pf_eligible=True, esic_eligible=True,
                        )])

                db.commit()
                result.success += 1
                result.rows.append(RowResult(row=row_idx, status="success", employee_code=emp_code, name=name))
            except Exception as exc:
                db.rollback()
                result.failed += 1
                result.rows.append(RowResult(row=row_idx, status="error", name=name, error=str(exc)))
                failed_raw.append((row_idx, raw, str(exc)))

    if failed_raw:
        result.failed_rows_xlsx_b64 = _build_failed_xlsx(failed_raw)

    result.rows.sort(key=lambda r: r.row)
    log.info("bulk_import: done — success=%d failed=%d", result.success, result.failed)
    return result
