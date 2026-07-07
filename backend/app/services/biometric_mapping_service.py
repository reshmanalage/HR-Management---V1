import re
from io import BytesIO

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.employee import Employee


def _parse_employee_header(cell_value: str):
    """'73 : Reshma Nalage' → ('73', 'Reshma Nalage')"""
    if not cell_value:
        return None, None
    m = re.match(r"^(\d+)\s*:\s*(.+)$", str(cell_value).strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return None, str(cell_value).strip()


def _normalise(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip()).lower()


def _find_employee(bio_name: str, name_to_emp: dict, employees) -> object | None:
    """
    Match a biometric display name to an Employee.
    Strategy 1 — exact full-name match (case-insensitive, normalised whitespace).
    Strategy 2 — for names with spaces: both first_name and last_name appear as
                 whole words in the bio name (guards against short substrings).
    Strategy 3 — for run-together names (no spaces): first_name AND last_name
                 are substrings, both at least 4 chars (avoids false positives).
    """
    norm = _normalise(bio_name)

    # Strategy 1: exact
    if norm in name_to_emp:
        return name_to_emp[norm]

    bio_words = set(norm.split())
    has_spaces = " " in norm

    best = None
    for emp in employees:
        fn = (emp.first_name or "").lower().strip()
        ln = (emp.last_name or "").lower().strip()
        if not fn or not ln:
            continue

        if has_spaces:
            # Both first and last must appear as whole words
            if fn in bio_words and ln in bio_words:
                best = emp
                break
        else:
            # Run-together name: substring match, minimum 4 chars each
            if len(fn) >= 4 and len(ln) >= 4 and fn in norm and ln in norm:
                best = emp
                break

    return best


def map_biometrics(file_bytes: bytes, db: Session) -> dict:
    """
    Parse a biometric attendance Excel, extract (biometric_code, name) pairs,
    match each to an employee in the DB by full name, and update biometric_code.
    Returns a report of matched, skipped (already set), and unmatched entries.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

    # Collect unique (code, name) pairs across all sheets
    bio_map: dict[str, str] = {}   # biometric_code → raw_name
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] != "Employee:":
                continue
            cell_val = row[3] if len(row) > 3 else None
            if not cell_val:
                continue
            code, name = _parse_employee_header(cell_val)
            if code and name:
                bio_map[code] = name

    if not bio_map:
        return {
            "matched": 0,
            "skipped": 0,
            "unmatched": [],
            "detail": [],
            "error": "No employee rows found in the uploaded file",
        }

    # Build normalised name → employee from DB
    employees = list(db.scalars(select(Employee)))
    name_to_emp: dict[str, Employee] = {
        _normalise(f"{e.first_name or ''} {e.last_name or ''}".strip()): e
        for e in employees
        if (e.first_name or e.last_name)
    }

    matched = 0
    skipped = 0
    unmatched: list[dict] = []
    detail: list[dict] = []

    for bio_code, raw_name in sorted(bio_map.items(), key=lambda x: int(x[0])):
        emp = _find_employee(raw_name, name_to_emp, employees)

        if emp is None:
            unmatched.append({"biometric_code": bio_code, "name": raw_name})
            continue

        if emp.biometric_code == bio_code:
            skipped += 1
            detail.append({
                "biometric_code": bio_code,
                "name": raw_name,
                "employee_code": emp.employee_code,
                "status": "already_set",
            })
            continue

        emp.biometric_code = bio_code
        db.add(emp)
        matched += 1
        detail.append({
            "biometric_code": bio_code,
            "name": raw_name,
            "employee_code": emp.employee_code,
            "status": "updated",
        })

    db.commit()

    return {
        "matched": matched,
        "skipped": skipped,
        "unmatched": unmatched,
        "detail": detail,
    }
