import re
from datetime import datetime
from io import BytesIO

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.attendance import AttendanceRecord
from app.models.employee import Employee


def _parse_duration_minutes(value) -> int | None:
    if not value:
        return None
    s = str(value).strip()
    if not s or s == "00:00":
        return None
    m = re.match(r"^(\d+):(\d{2})$", s)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    return None


def _parse_time(value) -> str | None:
    if not value:
        return None
    s = str(value).strip()
    if not s:
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})$", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return None


def _parse_cycle(text: str):
    """'May 21 2026  To  Jun 20 2026' → (date, date)"""
    parts = re.split(r"\s+To\s+", text.strip(), flags=re.IGNORECASE)
    if len(parts) != 2:
        return None, None
    try:
        start = datetime.strptime(parts[0].strip(), "%b %d %Y").date()
        end = datetime.strptime(parts[1].strip(), "%b %d %Y").date()
        return start, end
    except ValueError:
        return None, None


def _parse_employee_header(cell_value: str):
    """'73 : Reshma Nalage' → ('73', 'Reshma Nalage')"""
    if not cell_value:
        return None, None
    m = re.match(r"^(\d+)\s*:\s*(.+)$", str(cell_value).strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return None, str(cell_value).strip()


def import_attendance(file_bytes: bytes, db: Session) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

    # Pre-load employee code → id map
    emp_code_to_id: dict[str, int] = {
        row[0]: row[1]
        for row in db.execute(select(Employee.employee_code, Employee.id))
        if row[0]
    }

    inserted = 0
    skipped = 0
    errors: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find cycle period
        cycle_start = cycle_end = None
        for row in rows[:6]:
            for cell in row:
                if cell and isinstance(cell, str) and " To " in cell:
                    cycle_start, cycle_end = _parse_cycle(cell)
                    break
            if cycle_start:
                break

        if not cycle_start or not cycle_end:
            errors.append(f"Sheet '{sheet_name}': could not parse cycle dates")
            continue

        cycle_start_str = cycle_start.isoformat()
        cycle_end_str = cycle_end.isoformat()

        # Find Days header row and build col_idx → actual date map
        days_row = None
        for row in rows:
            if row and len(row) > 0 and row[0] == "Days":
                days_row = row
                break

        if days_row is None:
            errors.append(f"Sheet '{sheet_name}': could not find Days header row")
            continue

        col_date_map: dict[int, str] = {}
        for col_idx, cell in enumerate(days_row):
            if col_idx < 2 or not cell or not isinstance(cell, str):
                continue
            day_str = cell.strip().split()[0]
            if not day_str.isdigit():
                continue
            day_num = int(day_str)
            # Days >= cycle_start.day belong to start month; smaller days to end month
            if day_num >= cycle_start.day:
                d = cycle_start.replace(day=day_num)
            else:
                d = cycle_end.replace(day=day_num)
            col_date_map[col_idx] = d.isoformat()

        # Walk rows, detect employee blocks
        row_count = len(rows)
        i = 0
        while i < row_count:
            row = rows[i]
            if not row or not row[0] or row[0] != "Employee:":
                i += 1
                continue

            emp_code, emp_name = _parse_employee_header(row[3] if len(row) > 3 else None)
            if not emp_code:
                i += 1
                continue

            employee_id = emp_code_to_id.get(emp_code)

            # Grab the 8 data rows that follow
            def get_row(offset, label):
                idx = i + offset
                if idx < row_count and rows[idx] and rows[idx][0] == label:
                    return rows[idx]
                return None

            status_row = get_row(1, "Status")
            intime_row = get_row(2, "InTime")
            outtime_row = get_row(3, "OutTime")
            duration_row = get_row(4, "Duration")

            records_to_insert = []
            for col_idx, date_str in col_date_map.items():
                def _cell(r):
                    if r is None or col_idx >= len(r):
                        return None
                    return r[col_idx]

                status = _cell(status_row)
                in_t = _parse_time(_cell(intime_row))
                out_t = _parse_time(_cell(outtime_row))
                dur = _parse_duration_minutes(_cell(duration_row))

                records_to_insert.append(AttendanceRecord(
                    cycle_start=cycle_start_str,
                    cycle_end=cycle_end_str,
                    raw_employee_code=emp_code,
                    raw_employee_name=emp_name,
                    employee_id=employee_id,
                    date=date_str,
                    status=str(status) if status else None,
                    in_time=in_t,
                    out_time=out_t,
                    duration_minutes=dur,
                ))

            # Upsert: skip existing (cycle_start + emp_code + date)
            existing_dates = set(
                row[0]
                for row in db.execute(
                    select(AttendanceRecord.date).where(
                        AttendanceRecord.cycle_start == cycle_start_str,
                        AttendanceRecord.raw_employee_code == emp_code,
                    )
                )
            )

            for rec in records_to_insert:
                if rec.date in existing_dates:
                    skipped += 1
                else:
                    db.add(rec)
                    inserted += 1

            i += 9  # jump past the 8 data rows + employee header

    db.commit()
    return {"inserted": inserted, "skipped": skipped, "errors": errors}
